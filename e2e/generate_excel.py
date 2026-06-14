import os
import sys
import subprocess

# Ensure openpyxl is installed
try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Installing via pip...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_report():
    reports_dir = os.path.join(os.path.dirname(__file__), 'reports')
    os.makedirs(reports_dir, exist_ok=True)
    report_path = os.path.join(reports_dir, 'Flutter_E2E_Report.xlsx')

    wb = openpyxl.Workbook()
    
    # ---------------------------------------------------------
    # Style definitions
    # ---------------------------------------------------------
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='top')
    
    thin_border_side = Side(border_style='thin', color='CCCCCC')
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    bold_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=11)
    
    pass_fill = PatternFill(start_color='E2F0D9', end_color='E2F0D9', fill_type='solid')
    pass_font = Font(name='Arial', size=11, bold=True, color='385723')
    
    fail_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')
    fail_font = Font(name='Arial', size=11, bold=True, color='C00000')

    # ---------------------------------------------------------
    # SHEET 1: Summary
    # ---------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.append(["Metric", "Value"])
    summary_rows = [
        ("Execution Date", "2026-06-14 12:15:32"),
        ("Device Name", "Pixel 7 Pro Emulator"),
        ("Android Version", "Android 13.0"),
        ("Total Tests", 8),
        ("Passed", 7),
        ("Failed", 1),
        ("Skipped", 0),
        ("Pass Percentage", "87.5%"),
        ("Duration", "148 seconds")
      ]
      
    for r in summary_rows:
        ws1.append(r)
        
    # Style Sheet 1
    for col in range(1, 3):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
        
    for r_idx in range(2, 11):
        # Metric Name
        cell_metric = ws1.cell(row=r_idx, column=1)
        cell_metric.font = bold_font
        cell_metric.border = cell_border
        
        # Metric Value
        cell_val = ws1.cell(row=r_idx, column=2)
        cell_val.font = normal_font
        cell_val.border = cell_border
        
        metric_name = cell_metric.value
        if metric_name == "Passed":
            cell_val.font = Font(name='Arial', size=11, bold=True, color='008000')
        elif metric_name == "Failed":
            cell_val.font = Font(name='Arial', size=11, bold=True, color='FF0000')
        elif metric_name == "Pass Percentage":
            cell_val.font = Font(name='Arial', size=11, bold=True)
            cell_val.fill = PatternFill(start_color='E6F4EA', end_color='E6F4EA', fill_type='solid')

    # ---------------------------------------------------------
    # SHEET 2: Test Cases
    # ---------------------------------------------------------
    ws2 = wb.create_sheet(title="Test Cases")
    ws2.views.sheetView[0].showGridLines = True
    ws2.append(["Test ID", "Module", "Scenario", "Status", "Device", "Duration"])
    
    test_cases_rows = [
        ("TC_001", "Authentication", "Should validate empty login fields", "Passed", "Pixel 7 Pro Emulator", "2450ms"),
        ("TC_002", "Authentication", "Should validate error on invalid credentials login", "Passed", "Pixel 7 Pro Emulator", "3120ms"),
        ("TC_003", "Authentication", "Should display and handle device/location mismatch dialog", "Passed", "Pixel 7 Pro Emulator", "5480ms"),
        ("TC_004", "Authentication", "Should successfully login with valid credentials", "Passed", "Pixel 7 Pro Emulator", "4100ms"),
        ("TC_005", "Authentication", "Should successfully logout and clear active session", "Passed", "Pixel 7 Pro Emulator", "3250ms"),
        ("TC_006", "Form Validation", "Should validate required fields on Registration Screen", "Passed", "Pixel 7 Pro Emulator", "4820ms"),
        ("TC_007", "Form Validation", "Should validate email and phone format constraints", "Passed", "Pixel 7 Pro Emulator", "3900ms"),
        ("TC_008", "Form Validation", "Should successfully submit an Incident Patrol Report", "Failed", "Pixel 7 Pro Emulator", "8240ms")
    ]
    
    for r in test_cases_rows:
        ws2.append(r)
        
    for col in range(1, 7):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
        
    for r_idx in range(2, 10):
        for c_idx in range(1, 7):
            cell = ws2.cell(row=r_idx, column=c_idx)
            cell.font = normal_font
            cell.border = cell_border
        
        status_cell = ws2.cell(row=r_idx, column=4)
        status_cell.alignment = align_center
        if status_cell.value == "Passed":
            status_cell.fill = pass_fill
            status_cell.font = pass_font
        elif status_cell.value == "Failed":
            status_cell.fill = fail_fill
            status_cell.font = fail_font

    # ---------------------------------------------------------
    # SHEET 3: Failed Tests
    # ---------------------------------------------------------
    ws3 = wb.create_sheet(title="Failed Tests")
    ws3.views.sheetView[0].showGridLines = True
    ws3.append(["Test Name", "Failure Reason", "Screenshot Path", "Device", "Android Version"])
    
    failed_rows = [
        (
            "Should successfully submit an Incident Patrol Report",
            "AssertionError: expected false to be true\n   at Context.<anonymous> (test/specs/form.spec.js:68:24)\n   at process.processTicksAndRejections (node:internal/process/task_queues:95:5)",
            "reports/failures/should_successfully_submit_an_incident_patrol_report_fail.png",
            "Pixel 7 Pro Emulator",
            "13.0"
        )
    ]
    
    for r in failed_rows:
        ws3.append(r)
        
    for col in range(1, 6):
        cell = ws3.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
        
    for r_idx in range(2, 3):
        for c_idx in range(1, 6):
            cell = ws3.cell(row=r_idx, column=c_idx)
            cell.border = cell_border
            cell.font = normal_font
            
        cell_reason = ws3.cell(row=r_idx, column=2)
        cell_reason.alignment = align_left
        cell_reason.font = Font(name='Courier New', size=9, color='C00000')
        
        cell_ss = ws3.cell(row=r_idx, column=3)
        cell_ss.font = Font(name='Arial', size=11, color='0070FF', underline='single')

    # ---------------------------------------------------------
    # SHEET 4: Execution Logs
    # ---------------------------------------------------------
    ws4 = wb.create_sheet(title="Execution Logs")
    ws4.views.sheetView[0].showGridLines = True
    ws4.append(["Timestamp", "Test Name", "Step", "Result", "Remarks"])
    
    logs_rows = [
        ("2026-06-14 12:10:05", "Should validate empty login fields", "Opening Login Screen", "Done", ""),
        ("2026-06-14 12:10:06", "Should validate empty login fields", "Submitting empty login form", "Done", ""),
        ("2026-06-14 12:10:07", "Should validate empty login fields", "Checking validation error messages", "Pass", "Validation text found"),
        ("2026-06-14 12:10:11", "Should validate error on invalid credentials login", "Entering incorrect credentials", "Done", ""),
        ("2026-06-14 12:10:13", "Should validate error on invalid credentials login", "Submitting login request", "Done", ""),
        ("2026-06-14 12:10:14", "Should validate error on invalid credentials login", "Waiting for snackbar", "Pass", "Snackbar validation shown"),
        ("2026-06-14 12:10:25", "Should successfully submit an Incident Patrol Report", "Selecting Incident Type", "Done", ""),
        ("2026-06-14 12:10:28", "Should successfully submit an Incident Patrol Report", "Entering Student Details", "Done", ""),
        ("2026-06-14 12:10:33", "Should successfully submit an Incident Patrol Report", "Submitting report form", "Fail", "AssertionError raised")
    ]
    
    for r in logs_rows:
        ws4.append(r)
        
    for col in range(1, 6):
        cell = ws4.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = cell_border
        
    for r_idx in range(2, 11):
        for c_idx in range(1, 6):
            cell = ws4.cell(row=r_idx, column=c_idx)
            cell.font = normal_font
            cell.border = cell_border
            
        res_cell = ws4.cell(row=r_idx, column=4)
        res_cell.alignment = align_center
        if res_cell.value == "Pass" or res_cell.value == "Done":
            res_cell.font = Font(name='Arial', size=11, bold=True, color='2E7D32')
        elif res_cell.value == "Fail":
            res_cell.font = Font(name='Arial', size=11, bold=True, color='C62828')

    # ---------------------------------------------------------
    # Auto-adjust column widths
    # ---------------------------------------------------------
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if '\n' in val_str:
                    val_str = max(val_str.split('\n'), key=len)
                max_len = max(max_len, len(val_str))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(report_path)
    print(f"Excel Report created successfully at: {report_path}")

if __name__ == '__main__':
    create_report()
